import streamlit as st
import pandas as pd
import sqlite3
import hashlib
import io
import os
import re
import tempfile
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import msoffcrypto

# ============================================================
# 1. 页面设置
# ============================================================
st.set_page_config(page_title="物料提取工具", layout="wide")

# ============================================================
# 2. 万能使用码（已更新）
# ============================================================
MASTER_CODES = ["YVIP888", "Y1006"]

# ============================================================
# 3. 数据库
# ============================================================
def init_db():
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            remaining_uses INTEGER DEFAULT 3,
            is_permanent BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

def hash_pwd(pwd):
    return hashlib.sha256(pwd.encode()).hexdigest()

def get_user(username):
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("SELECT id, username, password, remaining_uses, is_permanent FROM users WHERE username = ?", (username,))
    user = c.fetchone()
    conn.close()
    return user

def create_user(username, password):
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    try:
        c.execute("INSERT INTO users (username, password, remaining_uses) VALUES (?, ?, 3)", (username, hash_pwd(password)))
        conn.commit()
        conn.close()
        return True
    except:
        conn.close()
        return False

def deduct_use(username):
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("UPDATE users SET remaining_uses = remaining_uses - 1 WHERE username = ? AND remaining_uses > 0", (username,))
    conn.commit()
    conn.close()

def add_permanent(username):
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("UPDATE users SET is_permanent = 1 WHERE username = ?", (username,))
    conn.commit()
    conn.close()


# ============================================================
# 4. 工具函数
# ============================================================
def is_encrypted(file_path):
    try:
        with open(file_path, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            return office_file.is_encrypted()
    except Exception as e:
        return False

def decrypt_file(file_path, password):
    decrypted_data = io.BytesIO()
    with open(file_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted_data)
    decrypted_data.seek(0)
    return decrypted_data

def read_recipe_table(decrypted_data):
    return pd.read_excel(decrypted_data, engine='openpyxl', header=None)

def process_merged_cells_from_bytes(decrypted_data):
    temp_file = "temp_decrypted.xlsx"
    with open(temp_file, "wb") as f:
        f.write(decrypted_data.getvalue())
    wb = load_workbook(temp_file, data_only=True)
    sheet_names = wb.sheetnames
    all_rows = []
    all_columns = set()
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        if not data or len(data) < 2:
            continue
        merged_ranges = list(ws.merged_cells.ranges)
        for merge_range in merged_ranges:
            min_row = merge_range.min_row
            min_col = merge_range.min_col
            max_row = merge_range.max_row
            max_col = merge_range.max_col
            top_left_value = data[min_row - 1][min_col - 1]
            for row in range(min_row - 1, max_row):
                for col in range(min_col - 1, max_col):
                    data[row][col] = top_left_value
        header = data[0]
        rows = data[1:]
        clean_header = []
        header_count = {}
        for col in header:
            col_str = str(col).strip() if col is not None else ''
            if col_str == '' or col_str == 'nan':
                col_str = f'Unnamed_{len(clean_header)}'
            if col_str in header_count:
                header_count[col_str] += 1
                col_str = f'{col_str}_{header_count[col_str]}'
            else:
                header_count[col_str] = 1
            clean_header.append(col_str)
            all_columns.add(col_str)
        for row in rows:
            if not row or all(v is None for v in row):
                continue
            row_dict = {'来源工作表': sheet_name}
            for i, col_name in enumerate(clean_header):
                if i < len(row):
                    value = row[i] if row[i] is not None else ''
                    if isinstance(value, str):
                        value = value.replace('\r\n', '\n').replace('\r', '\n')
                    row_dict[col_name] = value
                else:
                    row_dict[col_name] = ''
            all_rows.append(row_dict)
    if all_rows:
        all_columns = sorted(all_columns)
        all_columns_list = ['来源工作表'] + all_columns
        df_data = []
        for row_dict in all_rows:
            row_list = []
            for col in all_columns_list:
                row_list.append(row_dict.get(col, ''))
            df_data.append(row_list)
        df = pd.DataFrame(df_data, columns=all_columns_list)
        return df
    else:
        return pd.DataFrame()
    if os.path.exists(temp_file):
        try:
            os.remove(temp_file)
        except:
            pass

def find_column_by_keywords(df, keywords):
    for col in df.columns:
        col_str = str(col).strip()
        for keyword in keywords:
            if keyword in col_str:
                return col
    return None

def find_all_matching_columns(df, keywords, row_idx):
    if row_idx is None or row_idx >= len(df):
        return []
    matches = []
    for col_idx, val in enumerate(df.iloc[row_idx]):
        if pd.isna(val):
            continue
        val_str = str(val).strip()
        for keyword in keywords:
            if keyword in val_str:
                matches.append((col_idx, val_str))
                break
    return matches

def format_number(value):
    if value is None or value == '' or pd.isna(value):
        return ''
    try:
        num = float(value)
        if abs(num) < 0.000005:
            return 0.0
        return round(num, 5)
    except (ValueError, TypeError):
        return value

def apply_excel_formatting(file_path, title_text, red_rows):
    wb = load_workbook(file_path)
    ws = wb.active
    title_font = Font(name='宋体', size=18, bold=True)
    header_font = Font(name='宋体', size=11, bold=True)
    content_font = Font(name='宋体', size=11, bold=False)
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    red_font = Font(name='宋体', size=11, bold=False, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    max_col = ws.max_column
    max_row = ws.max_row
    header_row = 2 if title_text else 1
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = border
    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.alignment = center_alignment
        ws.row_dimensions[header_row].height = 30
    for row in range(header_row + 1, max_row + 1):
        ws.row_dimensions[row].height = None
        is_red = row in red_rows
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if is_red:
                cell.fill = red_fill
                cell.font = red_font
            else:
                cell.font = content_font
                cell.fill = PatternFill(fill_type=None)
            if col == 1:
                cell.alignment = center_alignment
            elif isinstance(cell.value, str) and '\n' in cell.value:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif col in [2, 3, 6, 7, 8, 9, 11, 12]:
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
    for col in range(1, max_col + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 60)
    wb.save(file_path)

def merge_group_rows(file_path, title_text, group_id):
    wb = load_workbook(file_path)
    ws = wb.active
    header_row = 2 if title_text else 1
    merge_cols = [1, 4, 7, 10, 11, 12]
    data_start_row = header_row + 1
    data_end_row = ws.max_row
    if data_start_row > data_end_row:
        wb.save(file_path)
        return
    row = data_start_row
    group_idx = 0
    while row <= data_end_row:
        if group_idx >= len(group_id):
            break
        current_group = group_id[group_idx]
        start_row = row
        group_count = 1
        while group_idx + group_count < len(group_id) and group_id[group_idx + group_count] == current_group:
            group_count += 1
        end_row = start_row + group_count - 1
        if group_count > 1:
            for col_idx in merge_cols:
                ws.merge_cells(start_row=start_row, start_column=col_idx,
                               end_row=end_row, end_column=col_idx)
                cell = ws.cell(row=start_row, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row = end_row + 1
        group_idx += group_count
    wb.save(file_path)

def get_column_letter_by_index(idx):
    result = ""
    while idx > 0:
        idx -= 1
        result = chr(65 + idx % 26) + result
        idx //= 26
    return result


# ============================================================
# 5. 核心提取函数
# ============================================================
def run_extraction(master_file, recipe_file, title_text, new_material_codes, master_pwd, recipe_pwd):
    try:
        # ---- 处理母表 ----
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            f.write(master_file.read())
            master_path = f.name
        master_file.seek(0)

        if master_pwd:
            decrypted_master = decrypt_file(master_path, master_pwd)
        else:
            with open(master_path, "rb") as f:
                decrypted_master = io.BytesIO(f.read())

        df_master = process_merged_cells_from_bytes(decrypted_master)
        if df_master.empty:
            os.unlink(master_path)
            return None

        id_col = find_column_by_keywords(df_master, ['编码', '编号'])
        if id_col is None:
            os.unlink(master_path)
            return None

        name_col = find_column_by_keywords(df_master, ['中文名称', '中文名'])
        inci_col = find_column_by_keywords(df_master, ['INCI', 'INCE'])
        content_col = find_column_by_keywords(df_master, ['含量', '金额'])
        report_col = find_column_by_keywords(df_master, ['报送码', '备案号', '注册号'])
        product_col = find_column_by_keywords(df_master, ['商品名称', '商品名'])
        supplier_col = find_column_by_keywords(df_master, ['生产商'])
        purpose_col = find_column_by_keywords(df_master, ['原料属性', '类别', '主要使用目的'])

        # ---- 处理配方表 ----
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            f.write(recipe_file.read())
            recipe_path = f.name
        recipe_file.seek(0)

        if recipe_pwd:
            decrypted_recipe = decrypt_file(recipe_path, recipe_pwd)
        else:
            with open(recipe_path, "rb") as f:
                decrypted_recipe = io.BytesIO(f.read())

        df_recipe_raw = read_recipe_table(decrypted_recipe)

        code_keywords = ['原料代码', '代码', '物料代码', '原料编号', '编号', '编码', 'code', 'Code']
        ratio_keywords = ['配比', '比例', '含量%', '添加量', '用量', '含量', 'ratio', 'Ratio']
        short_name_keywords = ['原料简称', '原料简名', '简名', '简称', '名称', '原料名', 'name', 'Name']

        header_row = None
        for idx, row in df_recipe_raw.iterrows():
            row_text = ' '.join([str(v) for v in row if pd.notna(v)])
            has_code = any(kw in row_text for kw in code_keywords)
            has_ratio = any(kw in row_text for kw in ratio_keywords)
            if has_code and has_ratio:
                header_row = idx
                break

        if header_row is None:
            for idx, row in df_recipe_raw.iterrows():
                row_text = ' '.join([str(v) for v in row if pd.notna(v)])
                if any(kw in row_text for kw in code_keywords):
                    header_row = idx
                    break

        if header_row is None:
            os.unlink(master_path)
            os.unlink(recipe_path)
            return None

        code_matches = find_all_matching_columns(df_recipe_raw, code_keywords, header_row)
        ratio_matches = find_all_matching_columns(df_recipe_raw, ratio_keywords, header_row)
        short_name_matches = find_all_matching_columns(df_recipe_raw, short_name_keywords, header_row)

        if len(code_matches) == 0 or len(ratio_matches) == 0:
            os.unlink(master_path)
            os.unlink(recipe_path)
            return None

        code_col = code_matches[0][0]
        ratio_col = ratio_matches[0][0]
        short_name_col = short_name_matches[0][0] if short_name_matches else None

        start_row = header_row + 1

        # ---- 阶段一：合并配比 ----
        raw_records = []
        for idx in range(start_row, len(df_recipe_raw)):
            code = df_recipe_raw.iloc[idx, code_col]
            ratio = df_recipe_raw.iloc[idx, ratio_col]
            has_code = pd.notna(code) and str(code).strip() != ''
            if not has_code:
                if short_name_col is not None:
                    sn = df_recipe_raw.iloc[idx, short_name_col]
                    if pd.notna(sn) and str(sn).strip():
                        raw_records.append({
                            'code': f"NOCODE_{sn}_{idx}",
                            'ratio': float(ratio) if pd.notna(ratio) else 0,
                            'short_name': str(sn).strip(),
                            'is_nocode': True
                        })
                continue
            code_str = str(code).strip()
            ratio_val = float(ratio) if pd.notna(ratio) else 0
            sn = ''
            if short_name_col is not None:
                sn_val = df_recipe_raw.iloc[idx, short_name_col]
                if pd.notna(sn_val):
                    sn = str(sn_val).strip()
            raw_records.append({
                'code': code_str,
                'ratio': ratio_val,
                'short_name': sn,
                'is_nocode': False
            })

        code_ratio_map = {}
        code_short_name_map = {}
        for rec in raw_records:
            if rec['is_nocode']:
                continue
            code = rec['code']
            if code not in code_ratio_map:
                code_ratio_map[code] = 0
                code_short_name_map[code] = rec['short_name']
            code_ratio_map[code] += rec['ratio']

        nocode_records = [rec for rec in raw_records if rec['is_nocode']]

        # ---- 阶段二：匹配母表生成结果 ----
        results = []
        group_id_list = []
        group_id_counter = 1
        is_new = False

        for code, total_ratio in code_ratio_map.items():
            match = df_master[df_master[id_col].astype(str).str.strip() == str(code).strip()]
            is_new = (new_material_codes is not None and code in new_material_codes)

            if match.empty:
                short_name = code_short_name_map.get(code, '')
                display_name = short_name if short_name else code
                results.append({
                    '原料序号': None,
                    '标准中文名称': display_name,
                    'INCI名称': '',
                    '原料含量 (%)': format_number(total_ratio),
                    '原料中成分含量 (%)': '',
                    '实际成分含量 (%)': '',
                    '主要使用目的': '',
                    '备注': '',
                    '是否新原料': '是' if is_new else '否',
                    '注册号/备案号': '',
                    '商品名': display_name,
                    '生产商': '',
                    '_code': code,
                    '_group_id': group_id_counter
                })
                group_id_list.append(group_id_counter)
                group_id_counter += 1
                continue

            first_row = match.iloc[0]
            purpose_val = first_row.get(purpose_col, '') if purpose_col else ''
            report_val = first_row.get(report_col, '') if report_col else ''
            product_val = first_row.get(product_col, '') if product_col else ''
            supplier_val = first_row.get(supplier_col, '') if supplier_col else ''

            for idx, master_row in match.iterrows():
                name_val = master_row.get(name_col, '') if name_col else ''
                if pd.isna(name_val):
                    name_val = ''
                if isinstance(name_val, str):
                    name_val = name_val.replace('\r\n', '\n').replace('\r', '\n')

                inci_val = master_row.get(inci_col, '') if inci_col else ''
                if pd.isna(inci_val):
                    inci_val = ''
                if isinstance(inci_val, str):
                    inci_val = inci_val.replace('\r\n', '\n').replace('\r', '\n')

                content_val = master_row.get(content_col, None) if content_col else None
                if pd.isna(content_val):
                    content_val = None
                else:
                    try:
                        content_val = float(content_val)
                    except:
                        content_val = None

                if content_val is not None:
                    actual_content = (total_ratio * content_val) / 100
                else:
                    actual_content = None

                results.append({
                    '原料序号': None,
                    '标准中文名称': name_val,
                    'INCI名称': inci_val,
                    '原料含量 (%)': format_number(total_ratio),
                    '原料中成分含量 (%)': format_number(content_val) if content_val is not None else '',
                    '实际成分含量 (%)': format_number(actual_content) if actual_content is not None else '',
                    '主要使用目的': purpose_val,
                    '备注': '',
                    '是否新原料': '是' if is_new else '否',
                    '注册号/备案号': report_val,
                    '商品名': product_val,
                    '生产商': supplier_val,
                    '_code': code,
                    '_group_id': group_id_counter
                })
                group_id_list.append(group_id_counter)

            group_id_counter += 1

        for rec in nocode_records:
            code = rec['code']
            sn = rec['short_name']
            ratio = rec['ratio']
            results.append({
                '原料序号': None,
                '标准中文名称': sn,
                'INCI名称': '',
                '原料含量 (%)': format_number(ratio),
                '原料中成分含量 (%)': '',
                '实际成分含量 (%)': '',
                '主要使用目的': '',
                '备注': '',
                '是否新原料': '否',
                '注册号/备案号': '',
                '商品名': sn,
                '生产商': '',
                '_code': code,
                '_group_id': group_id_counter
            })
            group_id_list.append(group_id_counter)
            group_id_counter += 1

        df_result = pd.DataFrame(results)

        # ---- 排序 ----
        df_result['_sort_key'] = pd.to_numeric(df_result['原料含量 (%)'], errors='coerce')
        df_result = df_result.sort_values(
            by=['_sort_key', '_group_id'],
            ascending=[False, True],
            na_position='last'
        )
        df_result = df_result.drop(columns=['_sort_key'])

        # ---- 生成原料序号 ----
        group_ids = df_result['_group_id'].unique()
        group_id_to_seq = {gid: idx + 1 for idx, gid in enumerate(group_ids)}
        df_result['原料序号'] = df_result['_group_id'].map(group_id_to_seq)
        sorted_group_ids = df_result['_group_id'].tolist()
        df_result = df_result.drop(columns=['_group_id', '_code'])

        # ---- 写入Excel ----
        base = title_text if title_text else "提取结果"
        output_file = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        df_result.to_excel(output_file, index=False)

        # ---- 插入标题行 ----
        if title_text:
            wb = load_workbook(output_file)
            ws = wb.active
            ws.insert_rows(1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
            cell = ws.cell(1, 1)
            cell.value = title_text
            cell.font = Font(name='宋体', size=18, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 40
            wb.save(output_file)

        # ---- 写入公式 ----
        wb = load_workbook(output_file)
        ws = wb.active
        header_row = 2 if title_text else 1
        col_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v:
                col_map[v] = c
        ratio_c = col_map.get('原料含量 (%)')
        content_c = col_map.get('原料中成分含量 (%)')
        actual_c = col_map.get('实际成分含量 (%)')

        if ratio_c and content_c and actual_c:
            r_letter = get_column_letter_by_index(ratio_c)
            c_letter = get_column_letter_by_index(content_c)
            group_first_row_map = {}
            data_start_row = header_row + 1
            row = data_start_row
            group_idx = 0
            while row <= ws.max_row and group_idx < len(sorted_group_ids):
                current_group = sorted_group_ids[group_idx]
                first_row_of_group = row
                group_count = 1
                while group_idx + group_count < len(sorted_group_ids) and sorted_group_ids[group_idx + group_count] == current_group:
                    group_count += 1
                for i in range(group_count):
                    group_first_row_map[row + i] = first_row_of_group
                row += group_count
                group_idx += group_count

            for row in range(header_row + 1, ws.max_row + 1):
                rv = ws.cell(row, ratio_c).value
                cv = ws.cell(row, content_c).value
                try:
                    if rv is not None and rv != '':
                        float(rv)
                        ws.cell(row, ratio_c).number_format = '0.00000'
                except:
                    pass
                try:
                    if cv is not None and cv != '':
                        float(cv)
                        ws.cell(row, content_c).number_format = '0.00000'
                except:
                    pass
                try:
                    if rv is not None and cv is not None and rv != '' and cv != '':
                        float(rv)
                        float(cv)
                        if row in group_first_row_map:
                            ratio_ref_row = group_first_row_map[row]
                        else:
                            ratio_ref_row = row
                        formula = f"={r_letter}{ratio_ref_row}*{c_letter}{row}/100"
                        cell = ws.cell(row, actual_c)
                        cell.value = formula
                        cell.number_format = '0.00000'
                except:
                    pass
            wb.save(output_file)

        # ---- 合并复合原料 ----
        merge_group_rows(output_file, title_text, sorted_group_ids)

        # ---- 标红逻辑 ----
        wb = load_workbook(output_file)
        ws = wb.active
        header_row = 2 if title_text else 1

        def find_col_by_keywords(ws, header_row, keywords):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(header_row, c).value
                if val:
                    for kw in keywords:
                        if kw in str(val):
                            return c
            return None

        name_c = find_col_by_keywords(ws, header_row, ['标准中文名称', '中文名称'])
        inci_c = find_col_by_keywords(ws, header_row, ['INCI名称', 'INCI'])
        content_c = find_col_by_keywords(ws, header_row, ['原料中成分含量', '含量'])
        purpose_c = find_col_by_keywords(ws, header_row, ['主要使用目的', '目的'])
        report_c = find_col_by_keywords(ws, header_row, ['注册号/备案号', '报送码', '备案号', '注册号'])
        product_c = find_col_by_keywords(ws, header_row, ['商品名', '商品名称'])
        supplier_c = find_col_by_keywords(ws, header_row, ['生产商'])

        required_cols = []
        col_names = []
        for col, name in zip([name_c, inci_c, content_c, purpose_c, report_c, product_c, supplier_c],
                               ['标准中文名称', 'INCI名称', '原料中成分含量', '主要使用目的', '注册号/备案号', '商品名', '生产商']):
            if col is not None:
                required_cols.append(col)
                col_names.append(name)

        data_rows = []
        for row in range(header_row + 1, ws.max_row + 1):
            row_data = []
            for c in range(1, ws.max_column + 1):
                row_data.append(ws.cell(row, c).value)
            data_rows.append((row, row_data))

        red_rows = set()
        row_index = 0

        while row_index < len(data_rows):
            excel_row, row_data = data_rows[row_index]
            current_seq = row_data[0]
            if current_seq is None or str(current_seq).strip() == '':
                row_index += 1
                continue

            group_rows = []
            while row_index < len(data_rows):
                next_seq = data_rows[row_index][1][0]
                if next_seq is not None and str(next_seq).strip() == str(current_seq).strip():
                    group_rows.append(data_rows[row_index])
                    row_index += 1
                else:
                    break

            if not group_rows:
                continue

            group_has_missing = False
            missing_cols_set = set()
            for excel_r, data in group_rows:
                for col_idx, col_name in zip(required_cols, col_names):
                    val = data[col_idx - 1]
                    if val is None or (isinstance(val, str) and val.strip() == ''):
                        group_has_missing = True
                        missing_cols_set.add(col_name)

            first_excel_row, first_data = group_rows[0]
            first_name = first_data[name_c - 1] if name_c is not None else ''
            first_seq = first_data[0]

            is_first_water = (first_seq is not None and str(first_seq).strip() == '1' and 
                              first_name is not None and str(first_name).strip() == '水')

            if is_first_water:
                pass
            elif group_has_missing:
                red_rows.add(first_excel_row)

        apply_excel_formatting(output_file, title_text, red_rows)

        # ---- 清理临时文件 ----
        os.unlink(master_path)
        os.unlink(recipe_path)

        return output_file

    except Exception as e:
        try:
            os.unlink(master_path)
        except:
            pass
        try:
            os.unlink(recipe_path)
        except:
            pass
        raise e


# ============================================================
# 6. 登录/注册界面
# ============================================================
def auth_page():
    st.title("🔐 登录 / 注册")
    st.caption("新用户自动获得 3 次免费使用机会")
    
    tab1, tab2 = st.tabs(["登录", "注册"])
    
    with tab1:
        username = st.text_input("用户名", key="login_user")
        password = st.text_input("密码", type="password", key="login_pwd")
        if st.button("登录"):
            user = get_user(username)
            if user and user[2] == hash_pwd(password):
                st.session_state.user = username
                st.rerun()
            else:
                st.error("用户名或密码错误")
    
    with tab2:
        new_user = st.text_input("设置用户名", key="reg_user")
        new_pwd = st.text_input("设置密码", type="password", key="reg_pwd")
        confirm_pwd = st.text_input("确认密码", type="password", key="reg_confirm")
        if st.button("注册"):
            if not new_user or not new_pwd:
                st.warning("请填写完整")
            elif new_pwd != confirm_pwd:
                st.warning("密码不一致")
            elif len(new_pwd) < 4:
                st.warning("密码至少4位")
            else:
                if create_user(new_user, new_pwd):
                    st.success("注册成功！请登录")
                else:
                    st.error("用户名已存在")
    
    st.markdown("---")
    st.caption("不想注册？")
    if st.button("👤 以游客身份体验（需付费解锁）"):
        st.session_state.user = "guest"
        st.rerun()


# ============================================================
# 7. 主功能界面
# ============================================================
def main_page():
    is_guest = (st.session_state.user == "guest")
    
    if is_guest:
        remaining = 0
        is_permanent = st.session_state.get("guest_authorized", False)
    else:
        user = get_user(st.session_state.user)
        if user is None:
            st.error("用户不存在，请重新登录")
            st.session_state.clear()
            st.rerun()
        remaining = user[3]
        is_permanent = user[4]
    
    st.title("📊 Excel物料提取工具")
    
    # 顶部状态栏
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        if is_guest:
            if is_permanent:
                st.success("✅ 游客 · 永久授权（已付费）")
            else:
                st.warning("⚠️ 游客模式 · 需付费解锁")
        else:
            if is_permanent:
                st.success("✅ 永久授权用户")
            else:
                st.info(f"📊 剩余免费次数：{remaining} 次")
    with col2:
        if st.button("🔑 万能码解锁"):
            st.session_state.show_master_input = True
    with col3:
        if st.button("🚪 退出"):
            st.session_state.clear()
            st.rerun()
    
    # 万能码输入弹窗
    if st.session_state.get("show_master_input", False):
        with st.expander("请输入万能使用码", expanded=True):
            master_code = st.text_input("万能使用码", type="password")
            if st.button("确认解锁"):
                if master_code.strip().upper() in [c.upper() for c in MASTER_CODES]:
                    if is_guest:
                        st.session_state.guest_authorized = True
                        st.success("🎉 游客授权成功！现在可以使用全部功能了")
                    else:
                        add_permanent(st.session_state.user)
                        st.success("🎉 解锁成功！已获得永久授权")
                    st.rerun()
                else:
                    st.error("❌ 使用码无效")
    
    st.markdown("---")
    
    # ====== 判断是否有权限使用 ======
    if not is_permanent:
        if is_guest:
            st.warning("⚠️ 游客模式需要付费解锁才能使用")
        else:
            if remaining <= 0:
                st.warning("⚠️ 您的免费次数已用完")
            else:
                # 有剩余次数，正常使用
                pass
        
        if (is_guest) or (not is_guest and remaining <= 0):
            st.markdown("""
            ### 💳 付费解锁永久授权
            
            请扫描下方二维码支付 **¥29.9**，然后输入上方的 **万能使用码** 即可永久解锁。
            
            （付费后联系客服获取使用码，或使用您已有的万能码）
            """)
            
            if os.path.exists("wechat_qr.png"):
                st.image("wechat_qr.png", caption="微信收款码", width=250)
            else:
                st.info("请将收款码图片命名为 wechat_qr.png 放在本程序同目录下")
            return
    
    # ====== 正常功能区域 ======
    st.subheader("📁 上传文件")
    
    title_text = st.text_input("表格标题（将作为文件名）", placeholder="例如：如微胶原抗皱面霜配方")
    
    master_file = st.file_uploader("上传母表", type=["xlsx", "xls"])
    recipe_file = st.file_uploader("上传配方表", type=["xlsx", "xls"])
    
    master_pwd = st.text_input("母表密码（如有）", type="password")
    recipe_pwd = st.text_input("配方表密码（如有）", type="password")
    
    new_material_option = st.radio("是否存在新原料？", ["否", "是"])
    new_material_codes = None
    if new_material_option == "是":
        codes = st.text_input("输入新原料代码（逗号分隔）")
        if codes:
            new_material_codes = [c.strip().upper() for c in codes.split(",") if c.strip()]
    
    if st.button("🚀 开始提取", type="primary"):
        if not master_file or not recipe_file:
            st.error("请上传母表和配方表")
        else:
            if not is_guest and not is_permanent:
                deduct_use(st.session_state.user)
            
            with st.spinner("⏳ 正在处理，请稍候..."):
                try:
                    result_file = run_extraction(
                        master_file, 
                        recipe_file, 
                        title_text, 
                        new_material_codes, 
                        master_pwd, 
                        recipe_pwd
                    )
                    
                    if result_file and os.path.exists(result_file):
                        with open(result_file, "rb") as f:
                            file_data = f.read()
                        
                        st.success("✅ 提取完成！")
                        st.download_button(
                            label="📥 下载结果表格",
                            data=file_data,
                            file_name=os.path.basename(result_file),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        os.unlink(result_file)
                    else:
                        st.error("❌ 处理失败，请检查文件格式是否正确（母表需包含'编码/编号'列，配方表需包含'原料代码'和'配比'列）")
                except Exception as e:
                    st.error(f"❌ 处理出错：{str(e)}")


# ============================================================
# 8. 程序入口
# ============================================================
def main():
    init_db()
    
    if "user" not in st.session_state:
        st.session_state.user = None
    if "show_master_input" not in st.session_state:
        st.session_state.show_master_input = False
    if "guest_authorized" not in st.session_state:
        st.session_state.guest_authorized = False
    
    # 路由
    if st.session_state.user:
        main_page()
    else:
        auth_page()

if __name__ == "__main__":
    main()
